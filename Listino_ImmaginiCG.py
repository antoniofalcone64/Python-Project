import openpyxl
import paramiko
import zipfile
import csv
import requests
import os
import sys

def porting_brand(riga_brand) :
    
    # apri il file XLSX
    workbook = openpyxl.load_workbook('REGOLE_BRAND.xlsx', read_only=True)

    # seleziona il foglio di lavoro
    sheet = workbook['Foglio1']

    # iterare su ogni riga del foglio di lavoro
    for riga in sheet.iter_rows():
        
        # riferimento alla cella "A1"
        cella_a = str(riga[0].value)
        cella_b = riga[1].value
        
        if(riga_brand == cella_a) :
            brand_tshop = cella_b
    
    return brand_tshop

sftp_path = "/Listino_#######.ZIP"
local_path = "./listino-zip/file.zip"
extract_path = "./listino-csv"

output_path = "./listino-xlsx-tshop"
extracted_csv_file = "./listino-csv/LISTINO_#######.CSV"
obsoleto = "N"
web = "SI"
iva = "22"
incr_image = 1
import_counter = 0
write = False
first_record = True

file_exist = True
output_file_name = '/output.xlsx'

incremental_name = 1
status = 0

excel_row = 2

# Connessione al server SFTP
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

try:
    ssh.connect('#######', username='#######', password='#######')
    print('Connessione SSH OK')

except (paramiko.AuthenticationException, paramiko.SSHException) as e:
    print(f"Connessione SSH FALLITA: {str(e)}")
    os.system('pause')
    sys.exit()


#fase preliminare 1.Controllo esistenza file output.xlsx
while(os.path.exists(output_path+output_file_name)) :
    output_file_name = '/output'+str(incremental_name)+'.xlsx'
    incremental_name = incremental_name + 1


# Crea un nuovo file XLSX
workbook = openpyxl.Workbook()

# Seleziona il foglio di lavoro
sheet = workbook.active


# Scaricamento del file .zip
try:
    sftp = ssh.open_sftp()
    print("Apertura canale SFTP OK")

except paramiko.SSHException as e:
    print(f"Apertura del canale SFTP FALLITA: {str(e)}")
    os.system('pause')
    sys.exit()

try:
    sftp.get(sftp_path, local_path)
    print("Download "+sftp_path+" OK\n")
    sftp.close()
except FileNotFoundError:
    print(f"Il file remoto {sftp_path} non esiste.\n")
    os.system('pause')
    sys.exit()
except Exception as e:
    print(f"Errore durante il download del file: {str(e)}\n")
    os.system('pause')
    sys.exit()


# Chiusura della connessione SSH
ssh.close()


# Estrazione del contenuto del file .zip
with zipfile.ZipFile(local_path, 'r') as zip_ref:
    zip_ref.extractall(extract_path)

#Recupero numero di righe file csv scaricato
num_row = 0
with open(extracted_csv_file, 'r') as csvrowfile:
    reader2 = csv.DictReader(csvrowfile, delimiter=';')
    for row in reader2:
        num_row = num_row + 1


#lettura file csv metodo chiavi dizionario
with open(extracted_csv_file, 'r') as csvfile:
    reader = csv.DictReader(csvfile, delimiter=';')
    for row in reader:
        categoria = " "


        #Sequenza di regole per filtraggio prodotti
        if(row['Descrizione Prodotto'] == 'ACCESS POINT' and (row['Descrizione Categoria'] == 'INDOOR' or row['Descrizione Categoria'] == 'OUTDOOR')) :
            categoria = '66'
            write = True
        
        if(row['Descrizione Prodotto'] == 'CASE') :
            categoria = '9'
            write = True

        if(row['Descrizione Prodotto'] == 'CAVETTERIA' and (row['Descrizione SottoCategoria'] == '50/125' or row['Descrizione SottoCategoria'] == '62.5' or row['Descrizione SottoCategoria'] == '9/125')) :
            categoria = '1156'
            write = True

        if(row['Descrizione Prodotto'] == 'CAVETTERIA' and (row['Descrizione Categoria'] == 'CAVI RETE' or row['Descrizione Categoria'] == 'CAVI RETE FIBRA' or row['Descrizione Categoria'] == 'CAVI FIBRA' or row['Descrizione Categoria'] == 'CAVI RAME' or row['Descrizione Categoria'] == 'CAVI ROUTER')) :
            categoria = '76'
            write = True

        if(row['Descrizione Prodotto'] == 'AUDIO') :
            categoria = '37'
            write = True

        if(row['Descrizione Prodotto'] == 'BITDEFENDER') :
            categoria = '99'
            write = True

        if(row['Descrizione Prodotto'] == 'CABLING') :
            categoria = '1156'
            write = True
        
        if(row['Descrizione Prodotto'] == 'HARD DISK' and row['Descrizione SottoCategoria'] == 'SATA 2.5') :
            categoria = '13'
            write = True

        if(row['Descrizione Prodotto'] == 'HARD DISK' and row['Descrizione SottoCategoria'] == 'SATA 3.5') :
            categoria = '14'
            write = True

        if(row['Descrizione Prodotto'] == 'HARD DISK' and (row['Descrizione SottoCategoria'] == 'SSD 3.5' or row['Descrizione SottoCategoria'] == 'SSD NAS' or row['Descrizione SottoCategoria'] == 'SSD ENTERPRISE' or row['Descrizione SottoCategoria'] == 'SSD SAS')) :
            categoria = '60'
            write = True

        if(row['Descrizione Prodotto'] == 'HARD DISK' and (row['Descrizione SottoCategoria'] == 'SSD' or row['Descrizione SottoCategoria'] == 'M.2 SATA' or row['Descrizione SottoCategoria'] == 'MSATA')) :
            categoria = '1000'
            write = True

        if(row['Descrizione Prodotto'] == 'LETTORI BARCODE') :
            categoria = '74'
            write = True

        if(row['Descrizione Prodotto'] == 'MAINBOARD') :
            categoria = '17'
            write = True

        if(row['Descrizione Prodotto'] == 'MEMORIE') :
            categoria = '15'
            write = True

        if(row['Descrizione Prodotto'] == 'MONITOR REFURBISHED') :
            categoria = '1142'
            write = True

        if(row['Descrizione Prodotto'] == 'MONITOR' and (row['Descrizione Categoria'] == 'OLED' or row['Descrizione Categoria'] == 'LED NO TOUCH')) :
            categoria = '86'
            write = True

        if(row['Descrizione Prodotto'] == 'NETWORKING' and row['Descrizione Categoria'] == 'TLEFONIA IP') :
            categoria = '73'
            write = True

        if(row['Descrizione Prodotto'] == 'NETWORKING' and (row['Descrizione Categoria'] == 'SWITCH' or row['Descrizione Categoria'] == 'SWITCH KVM')) :
            categoria = '270'
            write = True

        if(row['Descrizione Prodotto'] == 'NETWORKING' and row['Descrizione Categoria'] and (row['Descrizione SottoCategoria'] == 'CAVI IN RAME' or row['Descrizione SottoCategoria'] == 'CAVI IN FIBRA OTTICA' or row['Descrizione SottoCategoria'] == 'CAVI DI RETE FIBRA')) :
            categoria = '76'
            write = True

        if(row['Descrizione Prodotto'] == 'NOTEBOOK') :
            categoria = '63'
            write = True

        if(row['Descrizione Prodotto'] == 'NOTEBOOK REFURBISHED' or row['Descrizione Prodotto'] == 'PC REFURBISHED ALL-IN-ONE' or row['Descrizione Prodotto'] == 'PERSONAL COMPUTER REFURBISHED') :
            categoria = '1165'
            write = True

        if(row['Descrizione Prodotto'] == 'PERSONAL COMPUTER') :
            categoria = '34'
            write = True

        if(row['Descrizione Prodotto'] == 'POS') :
            categoria = '75'
            write = True
            
        if(row['Descrizione Prodotto'] == 'PLOTTER') :
            categoria = '1433'
            write = True
            
        if(row['Descrizione Prodotto'] == 'PROCESSORI' and row['Descrizione SottoCategoria'] != 'SOCKET SERVER') :
            categoria = '16'
            write = True

        if(row['Descrizione Prodotto'] == 'PROCESSORI' and row['Descrizione SottoCategoria'] == 'SOCKET SERVER') :
            categoria = '60'
            write = True

        if(row['Descrizione Prodotto'] == 'RAFFREDDAMENTO' and row['Descrizione Categoria'] == 'PERSONAL COMPUTER' and (row['Descrizione SottoCategoria'] == 'RAFFREDDAMENTO CPU A LIQUIDO' or row['Descrizione SottoCategoria'] == 'RAFFREDDAMENTO CPU AD ARIA')) :
            categoria = '18'
            write = True

        if(row['Descrizione Prodotto'] == 'RAFFREDDAMENTO' and row['Descrizione Categoria'] == 'PERSONAL COMPUTER' and row['Descrizione SottoCategoria'] == 'PASTA TERMICA') :
            categoria = '19'
            write = True

        if(row['Descrizione Prodotto'] == 'RAFFREDDAMENTO' and row['Descrizione Categoria'] == 'PERSONAL COMPUTER' and row['Descrizione SottoCategoria'] == 'VENTOLE PER CASE') :
            categoria = '20'
            write = True

        if(row['Descrizione Prodotto'] == 'ROUTER') :
            categoria = '67'
            write = True

        if(row['Descrizione Prodotto'] == 'SCHEDE GRAFICHE') :
            categoria = '11'
            write = True

        if(row['Descrizione Prodotto'] == 'SERVER') :
            categoria = '60'
            write = True

        if(row['Descrizione Prodotto'] == 'SWITCH') :
            categoria = '270'
            write = True

        if(row['Descrizione Prodotto'] == 'TABLET') :
            categoria = '65'
            write = True

        if(row['Descrizione Prodotto'] == 'TASTIERE E MOUSE') :
            categoria = '36'
            write = True
        
        if(row['Descrizione Prodotto'] == 'THIN CLIENT') :
            categoria = '34'
            write = True
        
        if(row['Descrizione Prodotto'] == 'TV/DIGITAL HOME/PROIETTORI' and row['Descrizione Categoria'] == 'VIDEOPROIETTORI') :
            categoria = '96'
            write = True

        if(row['Descrizione Prodotto'] == 'TV/DIGITAL HOME/PROIETTORI' and (row['Descrizione Categoria'] == 'APPLE TV' or row['Descrizione Categoria'] == 'HOTEL TV' or row['Descrizione Categoria'] == 'TV' or row['Descrizione Categoria'] == 'TV OLED')) :
            categoria = '94'
            write = True

        #inserimento riga articolo file xlsx finale
        if(first_record == True) :
            
            # Riga di intestazione
            sheet['A1'] = 'CODART'
            sheet['B1'] = 'DESART'
            sheet['C1'] = 'OBSOLETO'
            sheet['D1'] = 'CATEGORIA'
            sheet['E1'] = 'LISTINO9'
            sheet['F1'] = 'WEB'
            sheet['G1'] = 'IVA'
            sheet['H1'] = 'BRAND'

            first_record = False
            
        elif(write) :
            
            #download imagine articolo
            
            url = row['Immagine'] #url immagine da scaricare
            clean_id = row['Articolo Interno CG']
            clean_id = clean_id.replace('/', "-").replace(":", "-").replace("*", "-").replace("?", "-").replace('"', "-").replace("<", "-").replace(">", "-").replace("|", "-") 
            
            if(url != '') :
                try:   
                    response = requests.get(url, timeout=None)
                    filename = "C:/JOIA/t-shop/FOTO/ARTICOLI_OFFICE_ADOK/new/"+ clean_id +".jpg"
                    #filename = "./immagini/"+clean_id+".jpg"  
                    with open(filename, "wb") as f:
                        f.write(response.content)

                except requests.exceptions.RequestException as err:
                    print('ERRORE DOWNLOAD IMMAGINE: '+row['Immagine'])
                    filename =''
            else :
                filename=''    
            
           
            #scrittura righe articoli in file xlsx
        
            #print(clean_id, row['Descrizione Articolo'], obsoleto, categoria, row['Prezzo Vendita'], web, iva, porting_brand(row['Descrizione Produttore']))
            prezzo = float(row['Prezzo Vendita'].replace(",","."))
            prezzo = prezzo + ((prezzo/100)*10)
            sheet['A'+str(excel_row)] = clean_id
            sheet['B'+str(excel_row)] = row['Descrizione Articolo']
            sheet['C'+str(excel_row)] = obsoleto
            sheet['D'+str(excel_row)] = categoria
            sheet['E'+str(excel_row)] = prezzo
            sheet['F'+str(excel_row)] = web
            sheet['G'+str(excel_row)] = iva
            sheet['H'+str(excel_row)] = porting_brand(row['Descrizione Produttore'])

            print(clean_id, row['Descrizione Articolo'], obsoleto, categoria, row['Prezzo Vendita'], web, iva, sheet['H'+str(excel_row)].value)
            
            
            excel_row = excel_row + 1
           
            incr_image = incr_image + 1
            import_counter=import_counter + 1
            write = False
   
        
    # Salva il file XLSX
    workbook.save(output_path+output_file_name)
    
    
    print("Numero Articoli Totali: ",num_row)
    print("Numero Articoli Importati: ",import_counter)
    
    #os.system("start C:/Users/Antonio/Desktop/automazioneC.G/listino-csv-tshop"+output_file_name)
    print("C:/automazioneC.G/listino-csv-tshop/listino-xlsx-tshop"+output_file_name)
    
    
    os.system('pause')